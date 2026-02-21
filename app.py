"""
FIFA 25 Bot - Aplicação Flask Principal
Monitoramento de partidas do ESportsBattle
"""

import os
import logging
from datetime import datetime, timedelta, timezone
from io import BytesIO
import pytz
from flask import Flask, render_template, jsonify, request, send_file
from flask_sqlalchemy import SQLAlchemy
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
import atexit

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Inicialização do Flask
app = Flask(__name__)

# Configurações
app.config['SECRET_KEY'] = os.environ.get('SESSION_SECRET', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///fifa25.db')
if app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgres://'):
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

# Inicialização do banco de dados
db = SQLAlchemy(app)

# Timezone de Brasília
BRASILIA_TZ = pytz.timezone('America/Sao_Paulo')

def to_brasilia_time(dt):
    """Converte datetime UTC para horário de Brasília"""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(BRASILIA_TZ)

# Definir modelos inline para evitar import circular
class Match(db.Model):
    """Modelo de Partida"""
    __tablename__ = 'matches'
    
    id = db.Column(db.Integer, primary_key=True)
    match_id = db.Column(db.Integer, unique=True, nullable=False, index=True)
    status_id = db.Column(db.Integer, default=1)
    date = db.Column(db.DateTime, index=True)
    tournament_id = db.Column(db.Integer, index=True)
    tournament_token = db.Column(db.String(200))
    location_code = db.Column(db.String(100))
    location_name = db.Column(db.String(200))
    location_color = db.Column(db.String(20))
    console_id = db.Column(db.Integer)
    console_token = db.Column(db.String(100))
    player1_id = db.Column(db.Integer, index=True)
    player1_nickname = db.Column(db.String(100))
    player1_photo = db.Column(db.String(500))
    player1_team_id = db.Column(db.Integer)
    player1_team_name = db.Column(db.String(200))
    player1_team_logo = db.Column(db.String(500))
    player2_id = db.Column(db.Integer, index=True)
    player2_nickname = db.Column(db.String(100))
    player2_photo = db.Column(db.String(500))
    player2_team_id = db.Column(db.Integer)
    player2_team_name = db.Column(db.String(200))
    player2_team_logo = db.Column(db.String(500))
    score1 = db.Column(db.Integer)
    score2 = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    def to_dict(self):
        return {
            'id': self.id,
            'match_id': self.match_id,
            'status_id': self.status_id,
            'date': self.date.isoformat() if self.date else None,
            'player1_id': self.player1_id,
            'player1_nickname': self.player1_nickname or 'TBD',
            'player1_photo': self.player1_photo,
            'player1_team_id': self.player1_team_id,
            'player1_team_name': self.player1_team_name or 'N/A',
            'player1_team_logo': self.player1_team_logo,
            'player2_id': self.player2_id,
            'player2_nickname': self.player2_nickname or 'TBD',
            'player2_photo': self.player2_photo,
            'player2_team_id': self.player2_team_id,
            'player2_team_name': self.player2_team_name or 'N/A',
            'player2_team_logo': self.player2_team_logo,
            'score1': self.score1,
            'score2': self.score2,
            'location_code': self.location_code,
            'location_name': self.location_name or 'N/A',
            'location_color': self.location_color,
            'console_id': self.console_id,
            'console_token': self.console_token,
            'tournament_id': self.tournament_id,
            'tournament_token': self.tournament_token or 'N/A',
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }

class Player(db.Model):
    """Modelo de Jogador"""
    __tablename__ = 'players'
    
    id = db.Column(db.Integer, primary_key=True)
    player_id = db.Column(db.Integer, unique=True, nullable=False, index=True)
    nickname = db.Column(db.String(100), nullable=False)
    photo = db.Column(db.String(500))
    total_matches = db.Column(db.Integer, default=0)
    wins = db.Column(db.Integer, default=0)
    losses = db.Column(db.Integer, default=0)
    draws = db.Column(db.Integer, default=0)
    goals_scored = db.Column(db.Integer, default=0)
    goals_conceded = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

class Tournament(db.Model):
    """Modelo de Torneio"""
    __tablename__ = 'tournaments'
    
    id = db.Column(db.Integer, primary_key=True)
    tournament_id = db.Column(db.Integer, unique=True, nullable=False, index=True)
    status_id = db.Column(db.Integer, default=1)
    token = db.Column(db.String(200))
    token_international = db.Column(db.String(200))
    marker = db.Column(db.String(10))
    total_matches = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

class Analysis(db.Model):
    """Modelo de Análise Diária"""
    __tablename__ = 'analyses'
    
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, index=True)
    total_matches = db.Column(db.Integer, default=0)
    live_matches = db.Column(db.Integer, default=0)
    finished_matches = db.Column(db.Integer, default=0)
    canceled_matches = db.Column(db.Integer, default=0)
    unique_players = db.Column(db.Integer, default=0)
    top_teams = db.Column(db.Text)
    top_locations = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

# Importar serviços
from web_scraper import FIFA25Scraper
from data_analyzer import DataAnalyzer

try:
    from email_service import EmailService
    email_enabled = True
except:
    email_enabled = False
    logger.warning("⚠️ Email service não disponível")

try:
    from report_generator import ReportGenerator
    report_enabled = True
except:
    report_enabled = False
    logger.warning("⚠️ Report generator não disponível")

try:
    from telegram_service import TelegramService
    telegram_enabled = True
except:
    telegram_enabled = False
    logger.warning("⚠️ Telegram service não disponível")

# Variáveis globais
scraper = FIFA25Scraper()
analyzer = DataAnalyzer()
email_service = EmailService() if email_enabled else None
report_generator = ReportGenerator() if report_enabled else None
telegram = TelegramService() if telegram_enabled else None

# Configurações do scheduler
SCAN_INTERVAL = int(os.environ.get('SCAN_INTERVAL', 30))
RUN_SCRAPER = os.environ.get('RUN_SCRAPER', 'true').lower() == 'true'

# Estatísticas globais
stats = {
    'last_scan': None,
    'total_scans': 0,
    'total_matches': 0,
    'errors': 0,
    'status': 'Iniciando...',
    'success_rate': 100.0,
    'uptime': 0,
    'matches_per_hour': 0,
    'live_matches': 0,
    'upcoming_matches': 0,
    'finished_matches': 0,
    'unique_players': 0,
    'active_tournaments': 0,
    'avg_goals_per_match': 0,
    'most_active_player': 'N/A',
    'most_used_team': 'N/A',
    'busiest_location': 'N/A'
}

# Tempo de início do bot
bot_start_time = datetime.now()


def init_db():
    """Inicializa o banco de dados"""
    with app.app_context():
        try:
            db.create_all()
            logger.info("✅ Banco de dados inicializado")
        except Exception as e:
            logger.error(f"❌ Erro ao inicializar banco: {e}")


def run_scraper():
    """Executa o scraper de forma assíncrona"""
    if not RUN_SCRAPER:
        logger.info("⏸️ Scraper desabilitado (RUN_SCRAPER=false)")
        return
    
    with app.app_context():
        try:
            stats['status'] = 'Executando scraper...'
            logger.info("🔄 Iniciando varredura...")
            
            # 1. Buscar partidas próximas (endpoint principal)
            nearest_matches = scraper.get_nearest_matches()
            logger.info(f"📊 Encontradas {len(nearest_matches)} partidas próximas")
            
            # 2. Buscar partidas em streaming
            streaming_matches = scraper.get_streaming_matches()
            logger.info(f"📺 Encontradas {len(streaming_matches)} partidas em streaming")
            
            # 3. Processar e salvar no banco
            total_saved = 0
            
            # Processar nearest matches
            for match_data in nearest_matches:
                try:
                    match = save_match(match_data)
                    if match:
                        total_saved += 1
                except Exception as e:
                    logger.error(f"Erro ao salvar partida {match_data.get('id')}: {e}")
            
            # Processar streaming matches
            for match_data in streaming_matches:
                try:
                    match = save_match(match_data)
                    if match:
                        total_saved += 1
                except Exception as e:
                    logger.error(f"Erro ao salvar partida streaming {match_data.get('id')}: {e}")
            
            # 4. RE-BUSCAR partidas finalizadas para pegar placares atualizados
            finished_without_scores = Match.query.filter(
                Match.status_id == 3,
                db.or_(
                    Match.score1.is_(None),
                    Match.score2.is_(None)
                )
            ).limit(50).all()  # Limitar para não sobrecarregar
            
            if finished_without_scores:
                logger.info(f"🔄 Re-buscando {len(finished_without_scores)} partidas finalizadas sem placares...")
                
                for match in finished_without_scores:
                    try:
                        # Re-buscar a partida pela API
                        updated_match_data = scraper.get_match_by_id(match.match_id)
                        
                        if updated_match_data:
                            # Se agora tem placares, atualizar
                            if 'score1' in updated_match_data or 'participant1' in updated_match_data:
                                save_match(updated_match_data)
                                logger.info(f"✅ Placares atualizados para partida {match.match_id}")
                    except Exception as e:
                        logger.error(f"Erro ao re-buscar partida {match.match_id}: {e}")
            
            # 5. COLETAR ESTATÍSTICAS DOS TORNEIOS FINALIZADOS
            try:
                # Buscar torneios únicos das partidas finalizadas
                finished_tournaments = db.session.query(Match.tournament_id).filter(
                    Match.status_id == 3,
                    Match.tournament_id.isnot(None)
                ).distinct().limit(10).all()
                
                if finished_tournaments:
                    logger.info(f"📊 Coletando estatísticas de {len(finished_tournaments)} torneios...")
                    
                    for (tournament_id,) in finished_tournaments:
                        try:
                            # Buscar resultados do torneio
                            results_data = scraper.get_tournament_results(tournament_id)
                            
                            if results_data and 'results' in results_data:
                                results_list = results_data['results']
                                logger.info(f"✅ Torneio {tournament_id}: {len(results_list)} jogadores com estatísticas")
                                
                                # Aqui você pode salvar as estatísticas se quiser
                                # Por enquanto só logamos
                                for player_data in results_list:
                                    details = player_data.get('details', {})
                                    participant = player_data.get('participant', {})
                                    nickname = participant.get('nickname', 'N/A')
                                    
                                    logger.debug(f"  📈 {nickname}: {details.get('W')}V {details.get('L')}D {details.get('GF')}GF {details.get('GA')}GA")
                        except Exception as e:
                            logger.error(f"Erro ao coletar estatísticas do torneio {tournament_id}: {e}")
            except Exception as e:
                logger.error(f"Erro geral ao coletar estatísticas: {e}")
            
            # Atualizar estatísticas
            stats['last_scan'] = datetime.now()
            stats['total_scans'] += 1
            stats['total_matches'] = Match.query.count()
            stats['status'] = 'Online'
            
            # Calcular taxa de sucesso
            total_attempts = stats['total_scans']
            failures = stats['errors']
            if total_attempts > 0:
                stats['success_rate'] = round(((total_attempts - failures) / total_attempts) * 100, 1)
            else:
                stats['success_rate'] = 100.0
            
            # Calcular uptime (em horas)
            uptime_delta = datetime.now() - bot_start_time
            stats['uptime'] = round(uptime_delta.total_seconds() / 3600, 1)
            
            # Partidas por hora
            if stats['uptime'] > 0:
                stats['matches_per_hour'] = round(stats['total_matches'] / stats['uptime'], 1)
            
            # Estatísticas adicionais
            stats['live_matches'] = Match.query.filter_by(status_id=2).count()
            stats['upcoming_matches'] = Match.query.filter_by(status_id=1).count()
            stats['finished_matches'] = Match.query.filter_by(status_id=3).count()
            
            # Jogadores únicos
            unique_p1 = db.session.query(Match.player1_id).distinct().count()
            unique_p2 = db.session.query(Match.player2_id).distinct().count()
            stats['unique_players'] = unique_p1 + unique_p2
            
            # Torneios ativos
            stats['active_tournaments'] = db.session.query(Match.tournament_id).distinct().count()
            
            # Média de gols
            finished = Match.query.filter_by(status_id=3).filter(
                Match.score1.isnot(None),
                Match.score2.isnot(None)
            ).all()
            
            if finished:
                total_goals = sum([(m.score1 or 0) + (m.score2 or 0) for m in finished])
                stats['avg_goals_per_match'] = round(total_goals / len(finished), 2)
            
            # Jogador mais ativo
            top_player = db.session.query(
                Match.player1_nickname,
                db.func.count(Match.id).label('count')
            ).filter(
                Match.player1_nickname.isnot(None)
            ).group_by(
                Match.player1_nickname
            ).order_by(
                db.desc('count')
            ).first()
            
            if top_player:
                stats['most_active_player'] = top_player[0]
            
            # Time mais usado
            top_team = db.session.query(
                Match.player1_team_name,
                db.func.count(Match.id).label('count')
            ).filter(
                Match.player1_team_name.isnot(None)
            ).group_by(
                Match.player1_team_name
            ).order_by(
                db.desc('count')
            ).first()
            
            if top_team:
                stats['most_used_team'] = top_team[0]
            
            # Location mais ativa
            top_location = db.session.query(
                Match.location_name,
                db.func.count(Match.id).label('count')
            ).filter(
                Match.location_name.isnot(None)
            ).group_by(
                Match.location_name
            ).order_by(
                db.desc('count')
            ).first()
            
            if top_location:
                stats['busiest_location'] = top_location[0]
            
            logger.info(f"✅ Varredura completa: {total_saved} partidas salvas")
            
            # Enviar notificações se habilitado
            if telegram and total_saved > 0:
                try:
                    telegram.send_notification(f"🎮 {total_saved} novas partidas detectadas!")
                except:
                    pass
            
        except Exception as e:
            stats['errors'] += 1
            stats['status'] = f'Erro: {str(e)[:50]}'
            logger.error(f"❌ Erro no scraper: {e}")
            
            if telegram:
                try:
                    telegram.send_error(f"Erro no scraper: {e}")
                except:
                    pass


def send_weekly_report():
    """Envia relatório semanal por email"""
    if not email_enabled or not report_enabled:
        logger.warning("⚠️ Email ou Report Generator desabilitado")
        return
    
    with app.app_context():
        try:
            logger.info("📧 Gerando relatório semanal...")
            
            # Buscar partidas dos últimos 7 dias
            seven_days_ago = datetime.now() - timedelta(days=7)
            matches = Match.query.filter(
                Match.date >= seven_days_ago
            ).all()
            
            if not matches:
                logger.warning("⚠️ Nenhuma partida nos últimos 7 dias")
                return
            
            # Converter para lista de dicionários
            matches_data = [match.to_dict() for match in matches]
            
            # Gerar planilha Excel
            excel_path = report_generator.generate_weekly_report(matches_data)
            
            if not excel_path:
                logger.error("❌ Erro ao gerar planilha")
                return
            
            # Preparar dados do email
            total_matches = len(matches)
            finished = len([m for m in matches if m.status_id == 3])
            
            # Jogadores únicos
            players = set()
            for match in matches:
                if match.player1_nickname:
                    players.add(match.player1_nickname)
                if match.player2_nickname:
                    players.add(match.player2_nickname)
            
            report_data = {
                'total_matches': total_matches,
                'finished_matches': finished,
                'unique_players': len(players)
            }
            
            # Enviar email
            recipient_email = os.environ.get('RECIPIENT_EMAIL', os.environ.get('EMAIL_USER'))
            
            success = email_service.send_daily_report(
                to_address=recipient_email,
                report_data=report_data,
                attachment_path=excel_path
            )
            
            if success:
                logger.info(f"✅ Relatório semanal enviado para {recipient_email}")
            else:
                logger.error("❌ Falha ao enviar relatório semanal")
            
            # Limpar relatórios antigos
            report_generator.cleanup_old_reports(days=14)
            
        except Exception as e:
            logger.error(f"❌ Erro ao enviar relatório semanal: {e}")
            
            if email_service:
                try:
                    recipient_email = os.environ.get('RECIPIENT_EMAIL', os.environ.get('EMAIL_USER'))
                    email_service.send_error_notification(
                        to_address=recipient_email,
                        error_message=f"Erro ao gerar relatório semanal: {e}"
                    )
                except:
                    pass


def save_match(match_data):
    """Salva ou atualiza uma partida no banco de dados"""
    try:
        match_id = match_data.get('id')
        if not match_id:
            return None
        
        # Score - PEGAR DE DOIS LUGARES DIFERENTES!
        # Opção 1: score1 e score2 (nearest matches)
        score1 = match_data.get('score1')
        score2 = match_data.get('score2')
        
        # Opção 2: participant1.score e participant2.score (streaming matches)
        if score1 is None:
            p1_data = match_data.get('participant1', {})
            score1 = p1_data.get('score')
        
        if score2 is None:
            p2_data = match_data.get('participant2', {})
            score2 = p2_data.get('score')
        
        status_id = match_data.get('status_id', 1)
        
        logger.debug(f"💾 Salvando partida {match_id}: status={status_id}, score1={score1}, score2={score2}")
        
        # Verificar se já existe
        match = Match.query.filter_by(match_id=match_id).first()
        
        if not match:
            match = Match()
            match.match_id = match_id
        
        # Atualizar dados
        match.status_id = status_id
        match.date = datetime.fromisoformat(match_data.get('date', '').replace('Z', '+00:00')) if match_data.get('date') else None
        match.tournament_id = match_data.get('tournament_id')
        
        # Location
        location = match_data.get('location', {})
        match.location_code = location.get('code')
        match.location_name = location.get('token_international', location.get('token'))
        match.location_color = location.get('color')
        
        # Console
        console = match_data.get('console', {})
        match.console_id = console.get('id')
        match.console_token = console.get('token_international', console.get('token'))
        
        # Participant 1
        p1 = match_data.get('participant1', {})
        match.player1_id = p1.get('id')
        match.player1_nickname = p1.get('nickname')
        match.player1_photo = p1.get('photo')
        
        team1 = p1.get('team', {})
        match.player1_team_id = team1.get('id')
        match.player1_team_name = team1.get('token_international', team1.get('token'))
        match.player1_team_logo = team1.get('logo')
        
        # Participant 2
        p2 = match_data.get('participant2', {})
        match.player2_id = p2.get('id')
        match.player2_nickname = p2.get('nickname')
        match.player2_photo = p2.get('photo')
        
        team2 = p2.get('team', {})
        match.player2_team_id = team2.get('id')
        match.player2_team_name = team2.get('token_international', team2.get('token'))
        match.player2_team_logo = team2.get('logo')
        
        # Score - SALVAR OS PLACARES!
        match.score1 = score1
        match.score2 = score2
        
        # LOG se tem placar
        if score1 is not None and score2 is not None:
            logger.info(f"⚽ Partida {match_id}: {match.player1_nickname} {score1} x {score2} {match.player2_nickname}")
        
        # Tournament info
        tournament = match_data.get('tournament', {})
        match.tournament_token = tournament.get('token_international', tournament.get('token'))
        
        match.updated_at = datetime.now()
        
        db.session.merge(match)
        db.session.commit()
        
        return match
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao salvar partida: {e}")
        return None


# ==================== ROTAS ====================

@app.route('/')
def index():
    """Página inicial - Dashboard"""
    try:
        # Data atual
        today = datetime.now().date()
        
        # Buscar estatísticas gerais do banco
        total_matches = Match.query.count()
        live_matches_count = Match.query.filter_by(status_id=2).count()
        upcoming_matches_count = Match.query.filter_by(status_id=1).count()
        finished_matches_count = Match.query.filter_by(status_id=3).count()
        
        # Partidas do dia
        today_matches = Match.query.filter(
            db.func.date(Match.date) == today
        ).count()
        
        # Buscar partidas ao vivo AGORA
        live_matches_list = Match.query.filter_by(status_id=2).order_by(Match.date.desc()).limit(20).all()
        
        # Buscar próximas partidas agendadas
        upcoming_matches_list = Match.query.filter_by(status_id=1).order_by(Match.date.asc()).limit(20).all()
        
        # Buscar partidas finalizadas recentes
        finished_matches_list = Match.query.filter_by(status_id=3).order_by(Match.date.desc()).limit(10).all()
        
        # Estado da aplicação
        app_state = {
            'scheduler_running': RUN_SCRAPER,
            'email_enabled': email_enabled,
            'report_enabled': report_enabled,
            'database_connected': True,
            'last_error': None
        }
        
        # Preparar dados do summary
        summary = {
            'total_matches': total_matches,
            'today_matches': today_matches,
            'live_matches_count': live_matches_count,
            'upcoming_matches_count': upcoming_matches_count,
            'finished_matches_count': finished_matches_count,
            'nearest_matches_count': upcoming_matches_count,
            'recent_matches_count': finished_matches_count,
            'live_matches': [m.to_dict() for m in live_matches_list],
            'upcoming_matches': [m.to_dict() for m in upcoming_matches_list],
            'finished_matches': [m.to_dict() for m in finished_matches_list],
            'has_live_matches': live_matches_count > 0,
            'has_upcoming_matches': upcoming_matches_count > 0,
            'has_recent_matches': finished_matches_count > 0
        }
        
        # Garantir que stats tem TODOS os campos
        stats_copy = dict(stats)
        
        # Adicionar campos que podem estar faltando
        default_stats = {
            'last_scan': None,
            'total_scans': 0,
            'total_matches': 0,
            'errors': 0,
            'status': 'Iniciando...',
            'success_rate': 100.0,
            'uptime': 0,
            'matches_per_hour': 0,
            'live_matches': 0,
            'upcoming_matches': 0,
            'finished_matches': 0,
            'unique_players': 0,
            'active_tournaments': 0,
            'avg_goals_per_match': 0,
            'most_active_player': 'N/A',
            'most_used_team': 'N/A',
            'busiest_location': 'N/A',
            'scraper_enabled': RUN_SCRAPER,
            'scan_interval': SCAN_INTERVAL
        }
        
        # Mesclar defaults com valores atuais
        for key, default_value in default_stats.items():
            if key not in stats_copy or stats_copy[key] is None:
                stats_copy[key] = default_value
        
        # Formatar last_scan para string se existir
        if stats_copy.get('last_scan'):
            try:
                stats_copy['last_scan_formatted'] = stats_copy['last_scan'].strftime('%Y-%m-%d %H:%M:%S')
            except:
                stats_copy['last_scan_formatted'] = 'N/A'
        else:
            stats_copy['last_scan_formatted'] = 'Nunca'
        
        return render_template('dashboard.html', 
                             stats=stats_copy, 
                             summary=summary,
                             app_state=app_state)
        
    except Exception as e:
        logger.error(f"❌ Erro ao carregar dashboard: {e}")
        import traceback
        logger.error(traceback.format_exc())
        
        # Criar dados vazios para renderizar template sem erro
        empty_summary = {
            'total_matches': 0,
            'today_matches': 0,
            'live_matches_count': 0,
            'upcoming_matches_count': 0,
            'finished_matches_count': 0,
            'nearest_matches_count': 0,
            'recent_matches_count': 0,
            'live_matches': [],
            'upcoming_matches': [],
            'finished_matches': [],
            'has_live_matches': False,
            'has_upcoming_matches': False,
            'has_recent_matches': False
        }
        
        empty_stats = {
            'last_scan': None,
            'last_scan_formatted': 'Erro',
            'total_scans': 0,
            'total_matches': 0,
            'errors': 1,
            'status': 'Erro',
            'success_rate': 0.0,
            'uptime': 0,
            'matches_per_hour': 0,
            'live_matches': 0,
            'upcoming_matches': 0,
            'finished_matches': 0,
            'unique_players': 0,
            'active_tournaments': 0,
            'avg_goals_per_match': 0,
            'most_active_player': 'N/A',
            'most_used_team': 'N/A',
            'busiest_location': 'N/A',
            'scraper_enabled': False,
            'scan_interval': 30
        }
        
        empty_app_state = {
            'scheduler_running': False,
            'email_enabled': False,
            'report_enabled': False,
            'database_connected': False,
            'last_error': str(e)
        }
        
        # Tentar renderizar template com dados vazios
        try:
            return render_template('dashboard.html', 
                                 stats=empty_stats, 
                                 summary=empty_summary,
                                 app_state=empty_app_state)
        except Exception as template_error:
            logger.error(f"❌ Erro ao renderizar template: {template_error}")
            
            # Se até o template falhar, retornar HTML básico
            try:
                total = Match.query.count()
                live = Match.query.filter_by(status_id=2).count()
                upcoming = Match.query.filter_by(status_id=1).count()
            except:
                total = 0
                live = 0
                upcoming = 0
            
            return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>FIFA 25 Bot - Dashboard</title>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <style>
                * {{ margin: 0; padding: 0; box-sizing: border-box; }}
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    padding: 20px;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                }}
                .header {{
                    background: white;
                    padding: 30px;
                    border-radius: 15px;
                    box-shadow: 0 10px 30px rgba(0,0,0,0.2);
                    margin-bottom: 30px;
                    text-align: center;
                }}
                h1 {{ color: #667eea; font-size: 2.5em; margin-bottom: 10px; }}
                .status {{ color: #27ae60; font-size: 1.2em; }}
                .stats-grid {{
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                    gap: 20px;
                    margin-top: 30px;
                }}
                .stat-card {{
                    background: white;
                    padding: 25px;
                    border-radius: 15px;
                    box-shadow: 0 5px 15px rgba(0,0,0,0.1);
                    text-align: center;
                }}
                .stat-number {{
                    font-size: 3em;
                    font-weight: bold;
                    color: #667eea;
                    margin: 10px 0;
                }}
                .stat-label {{
                    color: #666;
                    font-size: 1.1em;
                }}
                .error-box {{
                    background: #fff3cd;
                    border-left: 4px solid #ffc107;
                    padding: 20px;
                    border-radius: 10px;
                    margin-top: 20px;
                }}
                .btn {{
                    display: inline-block;
                    padding: 15px 30px;
                    background: #667eea;
                    color: white;
                    text-decoration: none;
                    border-radius: 8px;
                    margin: 10px;
                    transition: all 0.3s;
                }}
                .btn:hover {{
                    background: #764ba2;
                    transform: translateY(-2px);
                    box-shadow: 0 5px 15px rgba(0,0,0,0.2);
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🎮 FIFA 25 Bot</h1>
                    <p class="status">✅ Sistema Online e Coletando Dados</p>
                </div>
                
                <div class="stats-grid">
                    <div class="stat-card">
                        <div class="stat-label">Total de Partidas</div>
                        <div class="stat-number">{total}</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-label">🔴 Ao Vivo</div>
                        <div class="stat-number">{live}</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-label">📅 Agendadas</div>
                        <div class="stat-number">{upcoming}</div>
                    </div>
                </div>
                
                <div class="error-box">
                    <h3>⚠️ Dashboard em Manutenção</h3>
                    <p>O template do dashboard precisa ser atualizado. Enquanto isso, use as APIs abaixo:</p>
                </div>
                
                <div style="text-align: center; margin-top: 30px;">
                    <a href="/api/stats" class="btn">📊 Ver Estatísticas (JSON)</a>
                    <a href="/api/matches/live" class="btn">🔴 Partidas Ao Vivo</a>
                    <a href="/api/matches/upcoming" class="btn">📅 Próximas Partidas</a>
                </div>
            </div>
        </body>
        </html>
        """, 500


    
    report_data = {
        'today_matches': today_matches,
        'live_matches': live_matches,
        'finished_today': finished_today,
        'total_matches': Match.query.count()
    }
    
    return render_template('reports.html', report=report_data)


# ==================== API ENDPOINTS ====================

@app.route('/api/stats')
def api_stats():
    """Retorna estatísticas do bot"""
    return jsonify({
        'last_scan': stats['last_scan'].isoformat() if stats['last_scan'] else None,
        'total_scans': stats['total_scans'],
        'total_matches': stats['total_matches'],
        'errors': stats['errors'],
        'status': stats['status'],
        'scraper_enabled': RUN_SCRAPER,
        'scan_interval': SCAN_INTERVAL
    })


@app.route('/api/matches/live')
def api_live_matches():
    """Retorna partidas ao vivo"""
    try:
        logger.info("🔴 API: Buscando partidas ao vivo...")
        matches = Match.query.filter_by(status_id=2).order_by(Match.date.desc()).limit(20).all()
        logger.info(f"🔴 API: Encontradas {len(matches)} partidas ao vivo")
        
        result = []
        for m in matches:
            try:
                result.append(m.to_dict())
            except Exception as e:
                logger.error(f"❌ Erro ao converter partida {m.match_id}: {e}")
                continue
        
        logger.info(f"🔴 API: Retornando {len(result)} partidas")
        return jsonify(result)
    
    except Exception as e:
        logger.error(f"❌ Erro na API /api/matches/live: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/matches/upcoming')
def api_upcoming_matches():
    """Retorna próximas partidas"""
    matches = Match.query.filter_by(status_id=1).order_by(Match.date.asc()).limit(20).all()
    return jsonify([m.to_dict() for m in matches])


@app.route('/api/matches/recent')
def api_recent_matches():
    """Retorna partidas recentes"""
    matches = Match.query.order_by(Match.updated_at.desc()).limit(50).all()
    return jsonify([m.to_dict() for m in matches])


@app.route('/api/force-scan')
def api_force_scan():
    """Força uma varredura imediata"""
    try:
        run_scraper()
        return jsonify({'success': True, 'message': 'Varredura iniciada'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/send-report')
def api_send_report():
    """Força envio de relatório semanal (para testes)"""
    try:
        send_weekly_report()
        return jsonify({'success': True, 'message': 'Relatório enviado'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== SCHEDULER ====================

def setup_scheduler():
    """Configura o scheduler para executar o scraper periodicamente"""
    scheduler = BackgroundScheduler()
    
    # Job 1: Scraper (a cada X segundos)
    scheduler.add_job(
        func=run_scraper,
        trigger=IntervalTrigger(seconds=SCAN_INTERVAL),
        id='scraper_job',
        name='Scraper FIFA25 ESportsBattle',
        replace_existing=True
    )
    logger.info(f"✅ Scheduler configurado: scraper a cada {SCAN_INTERVAL}s")
    
    # Job 2: Relatório Semanal (toda segunda-feira às 09:00)
    if email_enabled and report_enabled:
        from apscheduler.triggers.cron import CronTrigger
        
        scheduler.add_job(
            func=send_weekly_report,
            trigger=CronTrigger(day_of_week='mon', hour=9, minute=0),
            id='weekly_report_job',
            name='Relatório Semanal FIFA25',
            replace_existing=True
        )
        logger.info("✅ Scheduler configurado: relatório semanal toda segunda às 09:00")
    else:
        logger.warning("⚠️ Relatório semanal desabilitado (email ou report generator não disponível)")
    
    scheduler.start()
    
    # Desligar o scheduler quando a aplicação fechar
    atexit.register(lambda: scheduler.shutdown())
    
    return scheduler



# ==================== NOVAS ROTAS - ABAS COMPLETAS ====================

def calculate_player_stats(player_id):
    """Calcula estatísticas de um jogador"""
    matches = Match.query.filter(
        db.or_(
            Match.player1_id == player_id,
            Match.player2_id == player_id
        ),
        Match.status_id == 3
    ).all()
    
    if not matches:
        return None
    
    stats = {
        'player_id': player_id,
        'nickname': '',
        'total_matches': 0,
        'wins': 0,
        'losses': 0,
        'draws': 0,
        'goals_scored': 0,
        'goals_conceded': 0,
        'goal_difference': 0,
        'win_rate': 0.0
    }
    
    for match in matches:
        is_player1 = match.player1_id == player_id
        
        if is_player1 and not stats['nickname']:
            stats['nickname'] = match.player1_nickname or 'Unknown'
        elif not is_player1 and not stats['nickname']:
            stats['nickname'] = match.player2_nickname or 'Unknown'
        
        if match.score1 is not None and match.score2 is not None:
            stats['total_matches'] += 1
            
            if is_player1:
                stats['goals_scored'] += match.score1
                stats['goals_conceded'] += match.score2
                
                if match.score1 > match.score2:
                    stats['wins'] += 1
                elif match.score1 < match.score2:
                    stats['losses'] += 1
                else:
                    stats['draws'] += 1
            else:
                stats['goals_scored'] += match.score2
                stats['goals_conceded'] += match.score1
                
                if match.score2 > match.score1:
                    stats['wins'] += 1
                elif match.score2 < match.score1:
                    stats['losses'] += 1
                else:
                    stats['draws'] += 1
    
    stats['goal_difference'] = stats['goals_scored'] - stats['goals_conceded']
    
    if stats['total_matches'] > 0:
        stats['win_rate'] = round((stats['wins'] / stats['total_matches']) * 100, 1)
    
    return stats


@app.route('/matches')
def live_matches():
    """Página de partidas ao vivo"""
    try:
        logger.info("📄 ROTA /matches acessada")
        matches_list = Match.query.filter_by(status_id=2).order_by(Match.date.desc()).all()
        logger.info(f"📄 Encontradas {len(matches_list)} partidas ao vivo")
        
        for match in matches_list:
            if match.date:
                match.date_brasilia = to_brasilia_time(match.date).strftime('%d/%m/%Y %H:%M')
        
        logger.info(f"📄 Renderizando template matches.html com {len(matches_list)} partidas")
        return render_template('matches.html', matches=matches_list, total_matches=len(matches_list))
    
    except Exception as e:
        logger.error(f"❌ Erro na rota /matches: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return f"Erro ao carregar partidas: {str(e)}", 500


@app.route('/players')
def players_by_stadium():
    """Aba Jogadores - Players ativos por estádio"""
    try:
        # Buscar todos os jogadores únicos agrupados por estádio
        players_by_stadium = {}
        
        # Buscar todas as partidas
        matches = Match.query.all()
        
        for match in matches:
            stadium = match.location_name or 'Estádio Desconhecido'
            
            if stadium not in players_by_stadium:
                players_by_stadium[stadium] = set()
            
            if match.player1_nickname:
                players_by_stadium[stadium].add(match.player1_nickname)
            if match.player2_nickname:
                players_by_stadium[stadium].add(match.player2_nickname)
        
        # Converter sets para listas ordenadas
        players_by_stadium = {
            stadium: sorted(list(players)) 
            for stadium, players in sorted(players_by_stadium.items())
        }
        
        return render_template('players.html', players_by_stadium=players_by_stadium)
    
    except Exception as e:
        logger.error(f"❌ Erro na rota /players: {e}")
        return render_template('players.html', players_by_stadium={})


@app.route('/reports')
def reports_page():
    """Página de relatórios"""
    report_stats = {
        'total_matches': Match.query.count(),
        'finished_matches': Match.query.filter_by(status_id=3).count(),
        'live_matches': Match.query.filter_by(status_id=2).count(),
        'unique_players': 0
    }
    
    player_ids = set()
    all_matches = Match.query.all()
    for match in all_matches:
        if match.player1_id:
            player_ids.add(match.player1_id)
        if match.player2_id:
            player_ids.add(match.player2_id)
    
    report_stats['unique_players'] = len(player_ids)
    
    return render_template('reports.html', stats=report_stats)


@app.route('/history')
def history_recent():
    """Página de histórico - últimos 30 minutos"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = 20
        
        # Calcular 30 minutos atrás
        thirty_min_ago = datetime.now() - timedelta(minutes=30)
        
        # Query - apenas finalizadas dos últimos 30 minutos
        query = Match.query.filter(
            Match.status_id == 3,
            Match.updated_at >= thirty_min_ago
        )
        
        # Paginar
        pagination_obj = query.order_by(Match.updated_at.desc()).paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        matches_list = []
        
        # Converter para dicionários e formatar datas
        for match in pagination_obj.items:
            match_dict = {
                'match_id': match.match_id,
                'date': to_brasilia_time(match.date).strftime('%d/%m/%Y %H:%M') if match.date else 'N/A',
                'status_id': match.status_id,
                'player1_nickname': match.player1_nickname or 'TBD',
                'player1_team_name': match.player1_team_name or 'N/A',
                'player2_nickname': match.player2_nickname or 'TBD',
                'player2_team_name': match.player2_team_name or 'N/A',
                'score1': match.score1,
                'score2': match.score2,
                'location_name': match.location_name or 'N/A',
                'tournament_token': match.tournament_token or 'N/A'
            }
            matches_list.append(match_dict)
        
        pagination = {
            'page': page,
            'pages': pagination_obj.pages,
            'total': pagination_obj.total,
            'per_page': per_page,
            'has_prev': pagination_obj.has_prev,
            'has_next': pagination_obj.has_next,
            'prev_num': pagination_obj.prev_num,
            'next_num': pagination_obj.next_num
        }
        
        return render_template('history.html',
                             matches=matches_list,
                             pagination=pagination,
                             total_matches=pagination_obj.total,
                             date_from='',
                             date_to='',
                             player_filter='',
                             team_filter='')
    
    except Exception as e:
        logger.error(f"❌ Erro na rota /history: {e}")
        import traceback
        logger.error(traceback.format_exc())
        
        # Retornar página vazia em caso de erro
        return render_template('history.html',
                             matches=[],
                             pagination={'page': 1, 'pages': 1, 'total': 0, 'per_page': 20, 'has_prev': False, 'has_next': False, 'prev_num': None, 'next_num': None},
                             total_matches=0,
                             date_from='',
                             date_to='',
                             player_filter='',
                             team_filter='',
                             error=str(e))


@app.route('/statistics/test')
def statistics_test():
    """Teste de estatísticas - retorna JSON para debug"""
    try:
        logger.info("📊 TEST: Iniciando teste de estatísticas...")
        
        # Buscar partidas
        finished_matches = Match.query.filter_by(status_id=3).limit(10).all()
        logger.info(f"📊 TEST: {len(finished_matches)} partidas encontradas")
        
        result = {
            'total_matches': len(finished_matches),
            'samples': []
        }
        
        for match in finished_matches[:5]:
            result['samples'].append({
                'id': match.match_id,
                'player1': match.player1_nickname,
                'player2': match.player2_nickname,
                'score1': match.score1,
                'score2': match.score2,
                'stadium': match.location_name
            })
        
        return result
        
    except Exception as e:
        return {'error': str(e)}, 500


@app.route('/statistics')
def statistics():
    """Aba Estatísticas - Versão minimalista"""
    try:
        logger.info("=" * 80)
        logger.info("📊 INICIANDO ROTA /STATISTICS")
        logger.info("=" * 80)
        
        # Buscar partidas
        try:
            matches = Match.query.filter_by(status_id=3).all()
            logger.info(f"📊 Total de partidas finalizadas: {len(matches)}")
        except Exception as e:
            logger.error(f"❌ ERRO AO BUSCAR BANCO: {e}")
            return "Erro ao acessar banco de dados", 500
        
        if not matches:
            logger.info("📊 Nenhuma partida encontrada - retornando vazio")
            return render_template('statistics.html', stats_by_stadium={})
        
        # Processar
        stats = {}
        matches_processed = 0
        
        for match in matches:
            try:
                # Validação básica
                if not all([match.score1 is not None, match.score2 is not None, 
                           match.player1_nickname, match.player2_nickname]):
                    continue
                
                stadium = match.location_name or 'Desconhecido'
                
                if stadium not in stats:
                    stats[stadium] = {}
                
                # Player 1
                p1 = match.player1_nickname
                if p1 not in stats[stadium]:
                    stats[stadium][p1] = {'name': p1, 'wins': 0, 'losses': 0, 'draws': 0, 
                                          'goals_scored': 0, 'goals_conceded': 0, 'goal_diff': 0}
                
                stats[stadium][p1]['goals_scored'] += int(match.score1)
                stats[stadium][p1]['goals_conceded'] += int(match.score2)
                
                if match.score1 > match.score2:
                    stats[stadium][p1]['wins'] += 1
                elif match.score1 < match.score2:
                    stats[stadium][p1]['losses'] += 1
                else:
                    stats[stadium][p1]['draws'] += 1
                
                # Player 2
                p2 = match.player2_nickname
                if p2 not in stats[stadium]:
                    stats[stadium][p2] = {'name': p2, 'wins': 0, 'losses': 0, 'draws': 0,
                                          'goals_scored': 0, 'goals_conceded': 0, 'goal_diff': 0}
                
                stats[stadium][p2]['goals_scored'] += int(match.score2)
                stats[stadium][p2]['goals_conceded'] += int(match.score1)
                
                if match.score2 > match.score1:
                    stats[stadium][p2]['wins'] += 1
                elif match.score2 < match.score1:
                    stats[stadium][p2]['losses'] += 1
                else:
                    stats[stadium][p2]['draws'] += 1
                
                matches_processed += 1
                
            except Exception as e:
                logger.error(f"❌ Erro ao processar partida {match.match_id}: {e}")
                continue
        
        logger.info(f"📊 Partidas processadas: {matches_processed}")
        logger.info(f"📊 Estádios encontrados: {len(stats)}")
        
        # Calcular saldo
        for stadium in stats:
            for player in stats[stadium].values():
                player['goal_diff'] = player['goals_scored'] - player['goals_conceded']
        
        # Formatar para template
        result = {}
        for stadium in sorted(stats.keys()):
            players = sorted(stats[stadium].values(), key=lambda x: x['wins'], reverse=True)
            result[stadium] = {'players': players}
            logger.info(f"📊 {stadium}: {len(players)} jogadores")
        
        logger.info("📊 RENDERIZANDO TEMPLATE...")
        
        # Tentar renderizar
        try:
            return render_template('statistics.html', stats_by_stadium=result)
        except Exception as template_error:
            logger.error(f"❌ ERRO AO RENDERIZAR TEMPLATE: {template_error}")
            import traceback
            logger.error(traceback.format_exc())
            return f"Erro ao renderizar template: {str(template_error)}", 500
        
    except Exception as e:
        logger.error(f"❌ ERRO CRÍTICO: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return f"Erro crítico: {str(e)}", 500


@app.route('/upcoming')
def upcoming():
    """Página de partidas agendadas"""
    page = request.args.get('page', 1, type=int)
    per_page = 30
    
    query = Match.query.filter_by(status_id=1)
    
    pagination_obj = query.order_by(Match.date.asc()).paginate(page=page, per_page=per_page, error_out=False)
    
    matches_list = pagination_obj.items
    
    for match in matches_list:
        if match.date:
            match.date = to_brasilia_time(match.date).strftime('%d/%m/%Y %H:%M')
    
    pagination = {
        'page': page,
        'pages': pagination_obj.pages,
        'total': pagination_obj.total,
        'per_page': per_page,
        'has_prev': pagination_obj.has_prev,
        'has_next': pagination_obj.has_next,
        'prev_num': pagination_obj.prev_num,
        'next_num': pagination_obj.next_num
    }
    
    return render_template('upcoming.html', matches=matches_list, pagination=pagination, total_matches=pagination_obj.total)


def generate_excel_report(matches, filename):
    """Gera relatório Excel com formatação (verde/vermelho) separado por estádio"""
    import pandas as pd
    from openpyxl.styles import PatternFill
    
    # Se não há partidas, retornar planilha vazia
    if not matches:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_empty = pd.DataFrame({
                'Mensagem': ['Nenhuma partida com placares disponível no momento']
            })
            df_empty.to_excel(writer, sheet_name='Aviso', index=False)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{filename}.xlsx'
        )
    
    # Agrupar partidas por estádio (APENAS com placares válidos)
    matches_by_stadium = {}
    for match in matches:
        # PULAR partidas sem placares
        if match.score1 is None or match.score2 is None:
            continue
            
        stadium = match.location_name or 'Estádio Desconhecido'
        
        if stadium not in matches_by_stadium:
            matches_by_stadium[stadium] = []
        
        matches_by_stadium[stadium].append(match)
    
    # Se após filtrar não sobrou nada, retornar planilha vazia
    if not matches_by_stadium:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_empty = pd.DataFrame({
                'Mensagem': ['Nenhuma partida com placares disponível no momento. Aguarde as partidas finalizarem.']
            })
            df_empty.to_excel(writer, sheet_name='Aviso', index=False)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{filename}.xlsx'
        )
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Criar uma aba para cada estádio
        for stadium, stadium_matches in sorted(matches_by_stadium.items()):
            data = []
            
            for match in stadium_matches:
                # Determinar vencedor
                vencedor = ''
                if match.score1 > match.score2:
                    vencedor = match.player1_nickname or 'N/A'
                elif match.score2 > match.score1:
                    vencedor = match.player2_nickname or 'N/A'
                else:
                    vencedor = 'Empate'
                
                data.append({
                    'Data/Hora': to_brasilia_time(match.date).strftime('%d/%m/%Y %H:%M') if match.date else 'N/A',
                    'Jogador 1': match.player1_nickname or 'N/A',
                    'Time 1': match.player1_team_name or 'N/A',
                    'Gols P1': match.score1,
                    'Gols P2': match.score2,
                    'Jogador 2': match.player2_nickname or 'N/A',
                    'Time 2': match.player2_team_name or 'N/A',
                    'Torneio': match.tournament_token or 'N/A',
                    'Vencedor': vencedor
                })
            
            if data:
                df = pd.DataFrame(data)
                
                # Nome da aba (máximo 31 caracteres)
                sheet_name = stadium[:31] if len(stadium) > 31 else stadium
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formatação
                worksheet = writer.sheets[sheet_name]
                
                # Cores
                green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                
                # Aplicar formatação
                for row_idx, row_data in enumerate(data, start=2):
                    vencedor = row_data['Vencedor']
                    p1_name = row_data['Jogador 1']
                    p2_name = row_data['Jogador 2']
                    
                    # Empate - AMARELO para ambos
                    if vencedor == 'Empate':
                        worksheet[f'B{row_idx}'].fill = yellow_fill
                        worksheet[f'F{row_idx}'].fill = yellow_fill
                    else:
                        # Célula B (Jogador 1)
                        if vencedor == p1_name:
                            worksheet[f'B{row_idx}'].fill = green_fill
                        elif vencedor == p2_name:
                            worksheet[f'B{row_idx}'].fill = red_fill
                        
                        # Célula F (Jogador 2)
                        if vencedor == p2_name:
                            worksheet[f'F{row_idx}'].fill = green_fill
                        elif vencedor == p1_name:
                            worksheet[f'F{row_idx}'].fill = red_fill
                
                # Ajustar largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Adicionar aba de estatísticas
        stats_data = []
        stats_by_player = {}
        
        for match in matches:
            if match.player1_nickname and match.score1 is not None and match.score2 is not None:
                p1 = match.player1_nickname
                if p1 not in stats_by_player:
                    stats_by_player[p1] = {'wins': 0, 'losses': 0, 'goals_for': 0, 'goals_against': 0, 'championships': set()}
                
                stats_by_player[p1]['goals_for'] += match.score1
                stats_by_player[p1]['goals_against'] += match.score2
                
                if match.score1 > match.score2:
                    stats_by_player[p1]['wins'] += 1
                    if match.tournament_token:
                        stats_by_player[p1]['championships'].add(match.tournament_token)
                elif match.score1 < match.score2:
                    stats_by_player[p1]['losses'] += 1
            
            if match.player2_nickname and match.score1 is not None and match.score2 is not None:
                p2 = match.player2_nickname
                if p2 not in stats_by_player:
                    stats_by_player[p2] = {'wins': 0, 'losses': 0, 'goals_for': 0, 'goals_against': 0, 'championships': set()}
                
                stats_by_player[p2]['goals_for'] += match.score2
                stats_by_player[p2]['goals_against'] += match.score1
                
                if match.score2 > match.score1:
                    stats_by_player[p2]['wins'] += 1
                    if match.tournament_token:
                        stats_by_player[p2]['championships'].add(match.tournament_token)
                elif match.score2 < match.score1:
                    stats_by_player[p2]['losses'] += 1
        
        for player, stats in stats_by_player.items():
            stats_data.append({
                'Jogador': player,
                'Vitórias': stats['wins'],
                'Derrotas': stats['losses'],
                'Gols Marcados': stats['goals_for'],
                'Gols Sofridos': stats['goals_against'],
                'Saldo': stats['goals_for'] - stats['goals_against'],
                'Campeonatos': ', '.join(sorted(stats['championships'])) if stats['championships'] else '-'
            })
        
        if stats_data:
            df_stats = pd.DataFrame(stats_data)
            df_stats = df_stats.sort_values('Vitórias', ascending=False)
            df_stats.to_excel(writer, sheet_name='Estatísticas', index=False)
            
            worksheet_stats = writer.sheets['Estatísticas']
            for column in worksheet_stats.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet_stats.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'{filename}.xlsx')



@app.route('/api/download/all')
def download_all():
    """Download de todas as partidas finalizadas COM PLACARES"""
    matches = Match.query.filter(
        Match.status_id == 3,
        Match.score1.isnot(None),
        Match.score2.isnot(None)
    ).all()
    return generate_excel_report(matches, 'FIFA25_Todas_Partidas')


@app.route('/api/download/today')
def download_today():
    """Download das partidas finalizadas de hoje COM PLACARES"""
    today = datetime.now().date()
    matches = Match.query.filter(
        db.func.date(Match.date) == today,
        Match.status_id == 3,
        Match.score1.isnot(None),
        Match.score2.isnot(None)
    ).all()
    return generate_excel_report(matches, 'FIFA25_Partidas_Hoje')


@app.route('/api/download/custom')
def download_custom():
    """Download personalizado - apenas finalizadas COM PLACARES"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = Match.query.filter(
        Match.status_id == 3,
        Match.score1.isnot(None),
        Match.score2.isnot(None)
    )
    
    if date_from:
        from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        query = query.filter(db.func.date(Match.date) >= from_date)
    
    if date_to:
        to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        query = query.filter(db.func.date(Match.date) <= to_date)
    
    matches = query.all()
    filename = f'FIFA25_Personalizado_{date_from}_a_{date_to}'
    return generate_excel_report(matches, filename)


@app.route('/api/download/players')
def download_players():
    """Download estatísticas dos jogadores"""
    import pandas as pd
    
    players_data = []
    player_ids = set()
    
    all_matches = Match.query.filter_by(status_id=3).all()
    
    for match in all_matches:
        if match.player1_id:
            player_ids.add(match.player1_id)
        if match.player2_id:
            player_ids.add(match.player2_id)
    
    for player_id in player_ids:
        stats = calculate_player_stats(player_id)
        if stats and stats['total_matches'] > 0:
            players_data.append(stats)
    
    df = pd.DataFrame(players_data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Estatísticas', index=False)
    
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='FIFA25_Estatisticas_Jogadores.xlsx')


@app.route('/api/matches/count')
def matches_count():
    """Conta partidas por data"""
    date_str = request.args.get('date')
    
    if date_str:
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            count = Match.query.filter(db.func.date(Match.date) == date_obj).count()
            return jsonify({'count': count, 'date': date_str})
        except:
            return jsonify({'error': 'Data inválida'}), 400
    
    return jsonify({'error': 'Data não fornecida'}), 400

# ==================== INICIALIZAÇÃO ====================

# Inicializar banco de dados
init_db()

# Configurar scheduler
if RUN_SCRAPER:
    scheduler = setup_scheduler()
    logger.info("✅ Scheduler iniciado com sucesso")
else:
    logger.warning("⚠️ Scraper desabilitado")

# Executar primeira varredura ao iniciar
with app.app_context():
    run_scraper()


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    app.run(host='0.0.0.0', port=port, debug=debug)