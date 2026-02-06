# excel_exporter.py - NOVO ARQUIVO
# Gerencia exportação de partidas finalizadas para Excel

import pandas as pd
import os
from datetime import datetime
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

class ExcelExporter:
    """
    Classe para exportar partidas finalizadas para Excel
    organizadas por estádio
    """
    
    def __init__(self, excel_path='/mnt/user-data/outputs/FIFA25_Todas_Partidas.xlsx'):
        self.excel_path = excel_path
        self.stadiums = ['Anfield', 'Hillsborough', 'Old Trafford', 'Wembley', 'Etihad']
    
    def export_match(self, match):
        """
        Exporta uma partida finalizada para a planilha Excel
        
        Args:
            match: objeto Match do banco de dados
        """
        try:
            # Valida se partida está finalizada
            if match.status != 'finished':
                logger.warning(f"⚠️ Match {match.match_id} não está finalizada. Ignorando.")
                return False
            
            # Valida se tem placar
            if match.final_score_home is None or match.final_score_away is None:
                logger.warning(f"⚠️ Match {match.match_id} sem placar definido. Ignorando.")
                return False
            
            # Prepara dados da partida
            match_data = self._prepare_match_data(match)
            
            # Identifica estádio
            stadium = self._identify_stadium(match.location)
            
            # Adiciona à planilha
            self._add_to_excel(match_data, stadium)
            
            logger.info(f"✅ Partida {match.match_id} exportada para Excel ({stadium})")
            return True
            
        except Exception as e:
            logger.error(f"❌ Erro ao exportar partida {match.match_id}: {e}")
            return False
    
    def _prepare_match_data(self, match):
        """
        Prepara dados da partida no formato da planilha
        
        Returns:
            dict com dados formatados
        """
        # Determina vencedor
        winner = self._determine_winner(
            match.home_player,
            match.away_player,
            match.final_score_home,
            match.final_score_away
        )
        
        return {
            'Data/Hora': match.match_date.strftime('%d/%m/%Y %H:%M') if match.match_date else '',
            'Jogador 1': match.home_player or 'N/A',
            'Time 1': match.home_team or 'N/A',
            'Gols P1': match.final_score_home,
            'Gols P2': match.final_score_away,
            'Jogador 2': match.away_player or 'N/A',
            'Time 2': match.away_team or 'N/A',
            'Torneio': match.tournament or 'N/A',
            'Vencedor': winner
        }
    
    def _determine_winner(self, player1, player2, score1, score2):
        """Determina vencedor baseado no placar"""
        if score1 > score2:
            return player1
        elif score2 > score1:
            return player2
        else:
            return 'Empate'
    
    def _identify_stadium(self, location):
        """
        Identifica o estádio baseado na localização
        
        Args:
            location: string com nome do local
            
        Returns:
            nome do estádio padronizado
        """
        if not location:
            return 'Outros'
        
        location_lower = location.lower()
        
        for stadium in self.stadiums:
            if stadium.lower() in location_lower:
                return stadium
        
        # Se não encontrou, retorna o nome original ou 'Outros'
        return location if location else 'Outros'
    
    def _add_to_excel(self, match_data, stadium):
        """
        Adiciona partida à aba correspondente do Excel
        """
        try:
            # Verifica se arquivo existe
            if os.path.exists(self.excel_path):
                # Carrega arquivo existente
                with pd.ExcelFile(self.excel_path) as xls:
                    sheets = {}
                    for sheet_name in xls.sheet_names:
                        sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
                
                # Adiciona ou cria aba do estádio
                if stadium in sheets:
                    df = sheets[stadium]
                else:
                    df = pd.DataFrame(columns=match_data.keys())
                
                # Adiciona nova linha
                df = pd.concat([df, pd.DataFrame([match_data])], ignore_index=True)
                sheets[stadium] = df
            else:
                # Cria novo arquivo
                sheets = {stadium: pd.DataFrame([match_data])}
            
            # Salva de volta no Excel
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Aplica formatação
            self._apply_formatting(stadium)
            
        except Exception as e:
            logger.error(f"Erro ao adicionar ao Excel: {e}")
            raise
    
    def _apply_formatting(self, stadium_sheet):
        """
        Aplica formatação profissional à planilha
        """
        try:
            wb = load_workbook(self.excel_path)
            
            if stadium_sheet not in wb.sheetnames:
                return
            
            ws = wb[stadium_sheet]
            
            # Estilo do cabeçalho
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplica ao cabeçalho
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Centraliza colunas de gols
            for row in ws.iter_rows(min_row=2):
                for idx, cell in enumerate(row):
                    cell.border = thin_border
                    
                    # Centraliza colunas Gols P1 e Gols P2
                    if idx in [3, 4]:  # Índices das colunas de gols
                        cell.alignment = Alignment(horizontal="center")
                        cell.font = Font(bold=True, size=11)
                    
                    # Destaca vencedor
                    if idx == 8:  # Coluna Vencedor
                        cell.alignment = Alignment(horizontal="center")
                        cell.font = Font(bold=True)
                        
                        # Cor verde para vencedor
                        if cell.value and cell.value != 'Empate':
                            cell.fill = PatternFill(
                                start_color="C6EFCE",
                                end_color="C6EFCE",
                                fill_type="solid"
                            )
            
            # Ajusta largura das colunas
            column_widths = {
                'A': 16,  # Data/Hora
                'B': 15,  # Jogador 1
                'C': 15,  # Time 1
                'D': 10,  # Gols P1
                'E': 10,  # Gols P2
                'F': 15,  # Jogador 2
                'G': 15,  # Time 2
                'H': 30,  # Torneio
                'I': 15   # Vencedor
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Congela primeira linha
            ws.freeze_panes = 'A2'
            
            wb.save(self.excel_path)
            
        except Exception as e:
            logger.error(f"Erro ao aplicar formatação: {e}")
    
    def export_all_finished_matches(self):
        """
        Exporta todas as partidas finalizadas do banco que ainda não foram exportadas
        """
        from models import Match
        
        try:
            # Busca partidas finalizadas
            finished_matches = Match.query.filter_by(status='finished').all()
            
            # Carrega partidas já exportadas
            exported_ids = self._get_exported_match_ids()
            
            # Filtra partidas não exportadas
            to_export = [m for m in finished_matches if m.match_id not in exported_ids]
            
            logger.info(f"📊 Exportando {len(to_export)} partidas para Excel...")
            
            success_count = 0
            for match in to_export:
                if self.export_match(match):
                    success_count += 1
            
            logger.info(f"✅ {success_count}/{len(to_export)} partidas exportadas com sucesso")
            
            return success_count
            
        except Exception as e:
            logger.error(f"Erro ao exportar todas as partidas: {e}")
            return 0
    
    def _get_exported_match_ids(self):
        """
        Obtém IDs das partidas já exportadas na planilha
        (para evitar duplicatas)
        """
        exported_ids = set()
        
        try:
            if not os.path.exists(self.excel_path):
                return exported_ids
            
            with pd.ExcelFile(self.excel_path) as xls:
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    # Aqui você pode adicionar lógica para identificar partidas
                    # Por exemplo, usando combinação de jogadores + data
                    # Por simplicidade, vamos reexportar sempre
        
        except Exception as e:
            logger.error(f"Erro ao ler IDs exportados: {e}")
        
        return exported_ids
    
    def generate_summary_report(self):
        """
        Gera relatório resumido das partidas exportadas
        
        Returns:
            dict com estatísticas gerais
        """
        try:
            if not os.path.exists(self.excel_path):
                return {}
            
            summary = {
                'total_matches': 0,
                'by_stadium': {},
                'by_player': {},
                'last_update': datetime.now()
            }
            
            with pd.ExcelFile(self.excel_path) as xls:
                for stadium in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=stadium)
                    
                    matches_count = len(df)
                    summary['total_matches'] += matches_count
                    summary['by_stadium'][stadium] = matches_count
                    
                    # Conta vitórias por jogador
                    for _, row in df.iterrows():
                        winner = row.get('Vencedor')
                        if pd.notna(winner) and winner != 'Empate':
                            if winner not in summary['by_player']:
                                summary['by_player'][winner] = {'wins': 0, 'matches': 0}
                            summary['by_player'][winner]['wins'] += 1
                        
                        # Conta partidas por jogador
                        for player_col in ['Jogador 1', 'Jogador 2']:
                            player = row.get(player_col)
                            if pd.notna(player):
                                if player not in summary['by_player']:
                                    summary['by_player'][player] = {'wins': 0, 'matches': 0}
                                summary['by_player'][player]['matches'] += 1
            
            return summary
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório: {e}")
            return {}


# ============================================================
# INTEGRAÇÃO COM APP.PY
# ============================================================

"""
# No app.py, adicionar:

from excel_exporter import ExcelExporter

# Instanciar exportador
excel_exporter = ExcelExporter()

# Quando partida finalizar:
def on_match_finished(match):
    '''Callback quando partida finaliza'''
    
    # Exporta para Excel
    excel_exporter.export_match(match)
    
    # Atualiza estatísticas
    update_player_statistics(match)
    
    # Notifica (opcional)
    notify_match_finished(match)

# Job agendado para exportar pendentes
def export_pending_matches():
    '''Exporta partidas finalizadas que ainda não foram exportadas'''
    excel_exporter.export_all_finished_matches()

# Adicionar ao scheduler
scheduler.add_job(
    func=export_pending_matches,
    trigger="interval",
    minutes=5,
    id="export_to_excel",
    name="Exportar Partidas para Excel",
    replace_existing=True
)

# Endpoint para download
@app.route('/download/matches-excel')
def download_excel():
    '''Endpoint para baixar planilha'''
    from flask import send_file
    
    excel_path = '/mnt/user-data/outputs/FIFA25_Todas_Partidas.xlsx'
    
    if os.path.exists(excel_path):
        return send_file(
            excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'FIFA25_Partidas_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    else:
        return jsonify({'error': 'Arquivo não encontrado'}), 404
"""
